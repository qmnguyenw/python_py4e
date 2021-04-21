Python | Arithmetic operations in excel file using openpyxl



 **Prerequisite:** Reading & Writing to excel sheet using openpyxl

 **Openpyxl**is a Python library using which one can perform multiple
operations on excel files like reading, writing, arithmatic operations and
plotting graphs. Let’s see how to perform different arithmatic operations
using openpyxl.

  *  **=SUM(cell1:cell2) :** Adds all the numbers in a range of cells.

 __

 __  
 __

 __

 __  
 __  
 __

# import openpyxl module

import openpyxl

 

# Call a Workbook() function of openpyxl 

# to create a new blank Workbook object

wb = openpyxl.Workbook()

 

# Get workbook active sheet 

# from the active attribute.

sheet = wb.active

 

# writing to the cell of an excel sheet

sheet['A1'] = 200

sheet['A2'] = 300

sheet['A3'] = 400

sheet['A4'] = 500

sheet['A5'] = 600

 

# The value in cell A7 is set to a formula 

# that sums the values in A1, A2, A3, A4, A5 .

sheet['A7'] = '= SUM(A1:A5)'

 

# save the file

wb.save("sum.xlsx")  
  
---  
  
 __

 __

 **Output:**  
![](https://media.geeksforgeeks.org/wp-content/uploads/openpyxl-1.png)

  *  **=PRODUCT(cell1:cell2) :** Multiplies all the numbers in the range of cells.

 __

 __  
 __

 __

 __  
 __  
 __

import openpyxl

 

wb = openpyxl.Workbook()

sheet = wb.active

 

sheet['A1'] = 2

sheet['A2'] = 3

sheet['A3'] = 4

sheet['A4'] = 5

sheet['A5'] = 6

 

# The value in cell A7 is set to a formula 

# that multiplies the values in A1, A2, A3, A4, A5 .

sheet['A7'] = '= PRODUCT(A1:A5)'

 

wb.save("product.xlsx")  
  
---  
  
 __

 __

 **Output:**  
![](https://media.geeksforgeeks.org/wp-content/uploads/openpyxl2.png)

  *  **=AVERAGE(cell1:cell2) :** It gives the average (arithmetical mean) of all the numbers which is present in the given cell range.

 __

 __  
 __

 __

 __  
 __  
 __

import openpyxl

 

wb = openpyxl.Workbook()

sheet = wb.active

 

sheet['A1'] = 200

sheet['A2'] = 300

sheet['A3'] = 400

sheet['A4'] = 500

sheet['A5'] = 600

 

# The value in cell A7 is set to a formula 

# that return average of the values in A1, A2, A3, A4, A5 .

sheet['A7'] = '= AVERAGE(A1:A5)'

 

wb.save("average.xlsx")  
  
---  
  
 __

 __

 **Output:**  
![average](https://media.geeksforgeeks.org/wp-
content/uploads/Screenshot-1702-1.png)

  *  **=QUOTIENT(num1, num2) :** It returns the integer portion of a division.

 __

 __  
 __

 __

 __  
 __  
 __

import openpyxl

 

wb = openpyxl.Workbook()

sheet = wb.active

 

# The value in cell is set to a formula 

# that gives quotient value .

sheet['A1'] = '= QUOTIENT(64, 8)'

sheet['A2'] = '= QUOTIENT(25, 4)'

 

wb.save("quotient.xlsx")  
  
---  
  
 __

 __

 **Output:**  
![quotient](https://media.geeksforgeeks.org/wp-
content/uploads/Screenshot-1704-1.png)

  *  **=MOD(num1, num2) :** Returns the remainder after a number is divided by the divisor.

 __

 __  
 __

 __

 __  
 __  
 __

import openpyxl

 

wb = openpyxl.Workbook()

sheet = wb.active

 

# The value in cell is set to a formula 

# that gives remainder or modulus value.

sheet['A1'] = '= MOD(64, 8)'

sheet['A2'] = '= MOD(25, 4)'

 

wb.save("modulus.xlsx")  
  
---  
  
 __

 __

 **Output:**  
![modulus](https://media.geeksforgeeks.org/wp-
content/uploads/Screenshot-1705-1.png)

  *  **=COUNT(cell1:cell2) :** It counts the number of cells in a range that contain the number.

 __

 __  
 __

 __

 __  
 __  
 __

import openpyxl

 

wb = openpyxl.Workbook()

sheet = wb.active

 

sheet['A1'] = 200

sheet['A2'] = 300

sheet['A3'] = 400

sheet['A4'] = 500

sheet['A5'] = 600

 

# The value in cell A7 is set to a formula 

# that gives counting of number present in the cells.

sheet['A7'] = '= COUNT(A1:A6)'

 

wb.save("count.xlsx")  
  
---  
  
 __

 __

 **Output:**  
![](https://media.geeksforgeeks.org/wp-content/uploads/openpyxl_last.png)

Attention geek! Strengthen your foundations with the **Python Programming
Foundation** Course and learn the basics.

To begin with, your interview preparations Enhance your Data Structures
concepts with the **Python DS** Course.

  
  

  

My Personal Notes _arrow_drop_up_

Save

