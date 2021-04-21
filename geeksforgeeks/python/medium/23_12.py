Python | Writing to an excel file using openpyxl module



Prerequisite : Reading an excel file using openpyxl

 **Openpyxl** is a Python library for reading and writing Excel (with
extension xlsx/xlsm/xltx/xltm) files. The openpyxl module allows Python
program to read and modify Excel files.

For example, user might have to go through thousands of rows and pick out few
handful information to make small changes based on some criteria. Using
Openpyxl module, these tasks can be done very efficiently and easily.

Let’s see how to create and write to an excel-sheet using Python.

 **Code #1 :** Program to print a active sheet title name

  

  

 __

 __  
 __

 __

 __  
 __  
 __

# import openpyxl module

import openpyxl

 

# Call a Workbook() function of openpyxl 

# to create a new blank Workbook object

wb = openpyxl.Workbook()

 

# Get workbook active sheet 

# from the active attribute. 

sheet = wb.active

 

# Once have the Worksheet object,

# one can get its name from the

# title attribute.

sheet_title = sheet.title

 

print("active sheet title: " + sheet_title)  
  
---  
  
 __

 __

 **Output :**

    
    
    active sheet title: Sheet

  
**Code #2 :** Program to change the Title name

 __

 __  
 __

 __

 __  
 __  
 __

# import openpyxl module

import openpyxl

 

# Call a Workbook() function of openpyxl 

# to create a new blank Workbook object

wb = openpyxl.Workbook()

 

# Get workbook active sheet 

# from the active attribute

sheet = wb.active

 

# One can change the name of the title

sheet.title = "sheet1"

 

print("sheet name is renamed as: " + sheet.title)  
  
---  
  
 __

 __

 **Output :**

    
    
    sheet name is renamed as: sheet1

  
**Code #3 :** Program to write to an Excel sheet

 __

 __  
 __

 __

 __  
 __  
 __

# import openpyxl module

import openpyxl

 

# Call a Workbook() function of openpyxl 

# to create a new blank Workbook object

wb = openpyxl.Workbook()

 

# Get workbook active sheet 

# from the active attribute

sheet = wb.active

 

# Cell objects also have row, column

# and coordinate attributes that provide

# location information for the cell.

 

# Note: The first row or column integer

# is 1, not 0. Cell object is created by

# using sheet object's cell() method.

c1 = sheet.cell(row = 1, column = 1)

 

# writing values to cells

c1.value = "ANKIT"

 

c2 = sheet.cell(row= 1 , column = 2)

c2.value = "RAI"

 

# Once have a Worksheet object, one can

# access a cell object by its name also.

# A2 means column = 1 & row = 2.

c3 = sheet['A2']

c3.value = "RAHUL"

 

# B2 means column = 2 & row = 2.

c4 = sheet['B2']

c4.value = "RAI"

 

# Anytime you modify the Workbook object

# or its sheets and cells, the spreadsheet

# file will not be saved until you call

# the save() workbook method.

wb.save("C:\\Users\\user\\Desktop\\demo.xlsx")  
  
---  
  
 __

 __

 **Output :**  
![Output](https://media.geeksforgeeks.org/wp-
content/uploads/Screenshot-107.png)  

**code #4 :** Program to add Sheets in the Workbook

 __

 __  
 __

 __

 __  
 __  
 __

# import openpyxl module

import openpyxl

 

# Call a Workbook() function of openpyxl 

# to create a new blank Workbook object

wb = openpyxl.Workbook()

 

sheet = wb.active

 

# Sheets can be added to workbook with the

# workbook object's create_sheet() method. 

wb.create_sheet(index = 1 , title = "demo sheet2")

 

wb.save("C:\\Users\\user\\Desktop\\demo.xlsx")  
  
---  
  
 __

 __

 **Output :**  
![output](https://media.geeksforgeeks.org/wp-
content/uploads/Screenshot-108.png)

Attention geek! Strengthen your foundations with the **Python Programming
Foundation** Course and learn the basics.

To begin with, your interview preparations Enhance your Data Structures
concepts with the **Python DS** Course.

My Personal Notes _arrow_drop_up_

Save

